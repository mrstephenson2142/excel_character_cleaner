import pandas as pd
import re
import sys
import os
import unicodedata

def is_printable_character(char):
    """
    Determine if a character is printable/human-readable.
    
    Args:
        char: The character to check
        
    Returns:
        tuple: (is_printable, description)
    """
    if len(char) > 1:  # Handle escape sequences
        try:
            char = char.encode().decode('unicode_escape')
        except:
            return False, "Invalid escape sequence"
    
    code_point = ord(char)
    
    # Control characters are not printable
    if code_point < 32 or (code_point >= 127 and code_point < 160):
        return False, "Control character (non-printable)"
    
    # Check for Unicode category
    category = unicodedata.category(char)
    
    # Get character name if available
    try:
        name = unicodedata.name(char)
    except ValueError:
        name = "Unknown character"
    
    # Check if it's a defined Unicode character with a name
    if category.startswith('C'):  # Control, unassigned, private use etc.
        return False, f"Unicode category: {category} ({name})"
    
    # It's printable
    return True, f"Unicode: {name} (category: {category})"

def get_column_letter(col_idx):
    """
    Convert column index to Excel column letter (A, B, C, ..., Z, AA, AB, ...)
    """
    result = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result

def save_findings_to_text(file_path, results):
    """
    Save the detailed findings to a text file that matches console output.
    
    Args:
        file_path: Path to the original Excel file
        results: List of problematic character results
    
    Returns:
        Path to the text report file
    """
    if not results:
        return None
        
    report_file = f"{os.path.splitext(file_path)[0]}_findings_report.txt"
    
    try:
        with open(report_file, 'w', encoding='utf-8') as f:
            f.write(f"Problematic Character Report for: {os.path.basename(file_path)}\n")
            f.write(f"Generated on: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            f.write(f"Found {len(results)} instances of problematic characters:\n")
            f.write("-" * 80 + "\n")
            
            for res in results:
                f.write(f"Sheet: {res['sheet']}\n")
                f.write(f"Location: Cell {res['column']}{res['row']} (Column Header: {res['column_header']})\n")
                f.write(f"Problematic Character: {res['hex_value']}\n")
                
                if res['is_printable']:
                    f.write(f"Character: '{res['problematic_char']}' - {res['char_description']}\n")
                else:
                    f.write(f"Character: Non-printable - {res['char_description']}\n")
                    
                f.write(f"Character Position(s) in Cell: {res['char_positions']}\n")
                f.write(f"Cell Value: {res['cell_value']}\n")
                
                # If we have position information, also show the character context
                if 'char_positions_list' in res and res['char_positions_list']:
                    cell_str = str(res['cell_value'])
                    for pos in res['char_positions_list']:
                        # Create a visual pointer to the character position
                        context_start = max(0, pos - 10)
                        context_end = min(len(cell_str), pos + 11)
                        context = cell_str[context_start:context_end]
                        
                        # Calculate position of character in the context string
                        pointer_pos = pos - context_start
                        pointer = ' ' * pointer_pos + '^'
                        
                        f.write(f"Context: ...{context}...\n")
                        f.write(f"         {pointer}\n")
                
                f.write("-" * 80 + "\n")
                
        print(f"Detailed findings report saved to: {report_file}")
        return report_file
        
    except Exception as e:
        print(f"Error saving findings to text file: {e}")
        return None

def scan_excel_for_problematic_chars(file_path, problematic_chars=None):
    """
    Scan an Excel file for problematic characters and report their locations.
    
    Args:
        file_path: Path to the Excel file
        problematic_chars: List of problematic character patterns to search for (default: byte values 0x80-0xFF)
        
    Returns:
        List of tuples with (sheet_name, row, column, cell_value, matched_char)
    """
    if problematic_chars is None:
        # Default to searching for all extended ASCII characters (0x80-0xFF)
        problematic_chars = [chr(i) for i in range(0x80, 0x100)]
        # Add escape sequences that might appear in the string representation
        problematic_chars.extend([f"\\x{i:02x}" for i in range(0x80, 0x100)])
    
    results = []
    
    try:
        # Create Excel file reader
        excel_file = pd.ExcelFile(file_path)
        
        # Process each sheet
        for sheet_name in excel_file.sheet_names:
            print(f"Scanning sheet: {sheet_name}")
            
            # Read the sheet
            df = pd.read_excel(excel_file, sheet_name=sheet_name)
            
            # Scan each cell
            for row_idx, row in enumerate(df.itertuples(), 1):
                for col_idx, cell_value in enumerate(row[1:], 1):
                    # Only process string values
                    if isinstance(cell_value, str):
                        # Check for problematic characters
                        for char in problematic_chars:
                            if char in cell_value or re.search(f"{re.escape(char)}", str(cell_value)):
                                # Convert to column name (A, B, C, etc.)
                                col_name = get_column_letter(col_idx)
                                col_header = df.columns[col_idx-1]
                                
                                # Find positions of all occurrences of the character in the cell
                                positions = []
                                cell_str = str(cell_value)
                                
                                # For single characters
                                if len(char) == 1:
                                    start = 0
                                    while True:
                                        pos = cell_str.find(char, start)
                                        if pos == -1:
                                            break
                                        positions.append(pos)
                                        start = pos + 1
                                # For escape sequences like '\x81'
                                else:
                                    matches = list(re.finditer(re.escape(char), cell_str))
                                    positions = [m.start() for m in matches]
                                
                                # Format the positions as a list
                                positions_str = ", ".join([str(p) for p in positions])
                                
                                # Evaluate if the character is printable
                                is_printable, char_description = is_printable_character(char if len(char) == 1 else char.encode().decode('unicode_escape'))
                                
                                results.append({
                                    'sheet': sheet_name,
                                    'row': row_idx,
                                    'column': col_name,
                                    'column_header': col_header,
                                    'cell_value': cell_value,
                                    'problematic_char': char,
                                    'hex_value': f"0x{ord(char):02x}" if len(char) == 1 else char,
                                    'char_positions': positions_str,
                                    'char_positions_list': positions,
                                    'is_printable': is_printable,
                                    'char_description': char_description
                                })
    
    except Exception as e:
        print(f"Error scanning Excel file: {e}")
        return []
    
    return results

def clean_excel_file(file_path, results):
    """
    Clean problematic characters in an Excel file based on scanning results.
    
    Args:
        file_path: Path to the Excel file
        results: List of problematic character results from scan_excel_for_problematic_chars
    
    Returns:
        Path to the cleaned file
    """
    if not results:
        print("No problematic characters to clean.")
        return file_path
    
    try:
        # Create a copy of the Excel file
        base_name = os.path.splitext(file_path)[0]
        cleaned_file = f"{base_name}_cleaned.xlsx"
        
        # Keep track of the cleaned cells
        cleaned_cells = []
        
        # Group results by sheet for efficient processing
        results_by_sheet = {}
        for res in results:
            sheet = res['sheet']
            if sheet not in results_by_sheet:
                results_by_sheet[sheet] = []
            results_by_sheet[sheet].append(res)
        
        # Read the Excel file using openpyxl for direct cell manipulation
        import openpyxl
        workbook = openpyxl.load_workbook(file_path)
        
        for sheet_name, sheet_results in results_by_sheet.items():
            if sheet_name not in workbook.sheetnames:
                print(f"Warning: Sheet '{sheet_name}' not found in the workbook. Skipping.")
                continue
            
            worksheet = workbook[sheet_name]
            
            # Process each cell with problematic characters
            for res in sheet_results:
                # Convert from pandas 1-based indexing to openpyxl 1-based indexing
                # (both are 1-based for rows, but we need to convert column letter)
                cell_address = f"{res['column']}{res['row']}"
                cell = worksheet[cell_address]
                
                if cell.value is None:
                    print(f"Warning: Cell {cell_address} in sheet '{sheet_name}' is empty. Skipping.")
                    continue
                
                # Convert cell value to string if it's not already
                cell_value = str(cell.value)
                
                print(f"\nCleaning cell {sheet_name}!{cell_address}")
                print(f"Current value: {cell_value}")
                print(f"Problematic character: {res['hex_value']}")
                
                if res['is_printable']:
                    print(f"Character: '{res['problematic_char']}' - {res['char_description']}")
                else:
                    print(f"Character: Non-printable - {res['char_description']}")
                
                # Ask user what to do
                print("\nOptions:")
                print("1. Delete the character")
                print("2. Replace with custom text")
                print("3. Skip this cell")
                print("4. Skip all remaining cells")
                
                choice = input("Choose an option (1-4): ").strip()
                
                if choice == '4':
                    print("Skipping all remaining cells.")
                    break
                
                if choice == '3':
                    print(f"Skipping cell {cell_address}.")
                    continue
                
                if choice == '1':
                    # Delete the character
                    char_to_remove = res['problematic_char']
                    if len(char_to_remove) > 1:  # Handle escape sequences
                        try:
                            char_to_remove = char_to_remove.encode().decode('unicode_escape')
                        except:
                            print("Could not decode escape sequence. Skipping.")
                            continue
                    new_value = cell_value.replace(char_to_remove, '')
                    
                elif choice == '2':
                    # Replace with custom text
                    replacement = input("Enter replacement text: ")
                    char_to_replace = res['problematic_char']
                    if len(char_to_replace) > 1:  # Handle escape sequences
                        try:
                            char_to_replace = char_to_replace.encode().decode('unicode_escape')
                        except:
                            print("Could not decode escape sequence. Skipping.")
                            continue
                    new_value = cell_value.replace(char_to_replace, replacement)
                    
                else:
                    print("Invalid choice. Skipping this cell.")
                    continue
                
                # Update the cell value
                cell.value = new_value
                cleaned_cells.append({
                    'sheet': sheet_name,
                    'cell': cell_address,
                    'original': cell_value,
                    'cleaned': new_value
                })
                print(f"Updated cell value: {new_value}")
        
        # If any cells were cleaned, save the file
        if cleaned_cells:
            workbook.save(cleaned_file)
            print(f"\nCleaned {len(cleaned_cells)} cells.")
            print(f"Saved cleaned file as: {cleaned_file}")
            
            # Write summary of changes to a log file
            log_file = f"{base_name}_cleaning_log.txt"
            with open(log_file, 'w') as f:
                f.write(f"Excel Cleaning Log for {file_path}\n")
                f.write(f"Created on {pd.Timestamp.now()}\n\n")
                f.write(f"Total cells cleaned: {len(cleaned_cells)}\n\n")
                
                for i, cell_info in enumerate(cleaned_cells, 1):
                    f.write(f"Change {i}:\n")
                    f.write(f"  Sheet: {cell_info['sheet']}\n")
                    f.write(f"  Cell: {cell_info['cell']}\n")
                    f.write(f"  Original: {cell_info['original']}\n")
                    f.write(f"  Cleaned: {cell_info['cleaned']}\n\n")
                    
            print(f"Cleaning log saved as: {log_file}")
            return cleaned_file
        else:
            print("No changes were made.")
            return file_path
            
    except Exception as e:
        print(f"Error cleaning Excel file: {e}")
        return file_path

def main():
    if len(sys.argv) < 2:
        print("Usage: python excel_char_scanner.py <excel_file_path> [specific_chars]")
        print("Example to scan for \\x81: python excel_char_scanner.py myfile.xlsx \\x81")
        return
    
    file_path = sys.argv[1]
    
    if not os.path.exists(file_path):
        print(f"Error: File '{file_path}' not found.")
        return
    
    # If specific characters are provided, use them
    specific_chars = None
    if len(sys.argv) > 2:
        specific_chars = [sys.argv[2]]
        # Handle escape sequences properly
        specific_chars = [c.encode().decode('unicode_escape') if '\\' in c else c for c in specific_chars]
    
    results = scan_excel_for_problematic_chars(file_path, specific_chars)
    
    # Display results
    if not results:
        print(f"No problematic characters found in '{file_path}'.")
        return
    else:
        print(f"\nFound {len(results)} instances of problematic characters:")
        print("-" * 80)
        for res in results:
            print(f"Sheet: {res['sheet']}")
            print(f"Location: Cell {res['column']}{res['row']} (Column Header: {res['column_header']})")
            print(f"Problematic Character: {res['hex_value']}")
            if res['is_printable']:
                print(f"Character: '{res['problematic_char']}' - {res['char_description']}")
            else:
                print(f"Character: Non-printable - {res['char_description']}")
            print(f"Character Position(s) in Cell: {res['char_positions']}")
            print(f"Cell Value: {res['cell_value']}")
            
            # If we have position information, also show the character context
            if 'char_positions_list' in res and res['char_positions_list']:
                cell_str = str(res['cell_value'])
                for pos in res['char_positions_list']:
                    # Create a visual pointer to the character position
                    context_start = max(0, pos - 10)
                    context_end = min(len(cell_str), pos + 11)
                    context = cell_str[context_start:context_end]
                    
                    # Calculate position of character in the context string
                    pointer_pos = pos - context_start
                    pointer = ' ' * pointer_pos + '^'
                    
                    print(f"Context: ...{context}...")
                    print(f"         {pointer}")
            
            print("-" * 80)
            
    # Save results to CSV if there are any
    if results:
        # Save to CSV for technical users
        output_file = f"{os.path.splitext(file_path)[0]}_char_scan_results.csv"
        pd.DataFrame(results).to_csv(output_file, index=False)
        print(f"Results saved to {output_file}")
        
        # Save detailed findings to text file for less technical users
        report_file = save_findings_to_text(file_path, results)
        
        # Ask if user wants to clean the file
        print("\nWould you like to clean the problematic characters?")
        print("This will create a new copy of the Excel file with the problematic characters handled.")
        choice = input("Clean the file? (y/n): ").strip().lower()
        
        if choice == 'y' or choice == 'yes':
            cleaned_file = clean_excel_file(file_path, results)
            if cleaned_file != file_path:
                print(f"\nCleaning complete! You can now use the cleaned file: {cleaned_file}")
        else:
            print("No cleaning performed. You can manually edit the file using the scan results.")

if __name__ == "__main__":
    main()
